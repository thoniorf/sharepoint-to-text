"""
XLSX Spreadsheet Extractor

Extracts text content and metadata from Microsoft Excel .xlsx files
(Office Open XML format, Excel 2007+).

Uses openpyxl for parsing cells/sheets and direct ZIP parsing for images.
"""

import datetime
import io
import logging
from typing import Any, Generator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from sharepoint2text.exceptions import (
    ExtractionError,
    ExtractionFailedError,
    ExtractionFileEncryptedError,
)
from sharepoint2text.extractors.data_types import (
    XlsxContent,
    XlsxImage,
    XlsxMetadata,
    XlsxSheet,
)
from sharepoint2text.extractors.util.encryption import is_ooxml_encrypted
from sharepoint2text.extractors.util.ooxml_context import OOXMLZipContext
from sharepoint2text.extractors.util.zip_bomb import validate_zip_bytesio
from sharepoint2text.extractors.util.zip_utils import parse_relationships

logger = logging.getLogger(__name__)

# =============================================================================
# XML Namespaces and pre-computed tag names
# =============================================================================

XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

XDR_ONE_CELL_ANCHOR = f"{{{XDR_NS}}}oneCellAnchor"
XDR_TWO_CELL_ANCHOR = f"{{{XDR_NS}}}twoCellAnchor"
XDR_ABSOLUTE_ANCHOR = f"{{{XDR_NS}}}absoluteAnchor"
XDR_PIC = f"{{{XDR_NS}}}pic"
XDR_EXT = f"{{{XDR_NS}}}ext"
XDR_NVPICPR = f"{{{XDR_NS}}}nvPicPr"
XDR_CNVPR = f"{{{XDR_NS}}}cNvPr"
XDR_BLIPFILL = f"{{{XDR_NS}}}blipFill"
A_BLIP = f"{{{A_NS}}}blip"
R_EMBED = f"{{{R_NS}}}embed"

ANCHOR_TYPES = (XDR_ONE_CELL_ANCHOR, XDR_TWO_CELL_ANCHOR, XDR_ABSOLUTE_ANCHOR)

# =============================================================================
# Constants
# =============================================================================

EMU_PER_PIXEL = 9525

_CONTENT_TYPE_MAP = {
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "gif": "image/gif",
    "bmp": "image/bmp",
    "tiff": "image/tiff",
    "tif": "image/tiff",
    "emf": "image/x-emf",
    "wmf": "image/x-wmf",
}

# JPEG SOF markers for dimension extraction
_JPEG_SOF_MARKERS = frozenset(
    {
        0xC0,
        0xC1,
        0xC2,
        0xC3,
        0xC5,
        0xC6,
        0xC7,
        0xC9,
        0xCA,
        0xCB,
        0xCD,
        0xCE,
        0xCF,
    }
)

# Datetime types for isinstance check
_DATETIME_TYPES = (datetime.datetime, datetime.date, datetime.time)


# =============================================================================
# Cell value handling
# =============================================================================


def _get_cell_value(cell_value: Any) -> Any:
    """Convert cell value to appropriate Python type (datetime -> ISO string)."""
    if cell_value is None:
        return None
    if isinstance(cell_value, _DATETIME_TYPES):
        return cell_value.isoformat()
    return cell_value


def _format_value_for_display(value: Any) -> str:
    """Format a value as string for text table display."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _is_cell_non_empty(val: Any) -> bool:
    """Check if a cell value is non-empty."""
    return val is not None and (not isinstance(val, str) or val.strip() != "")


def _is_meaningful_value(val: Any) -> bool:
    """Check if value is meaningful (not None, empty, or 'Unnamed:' placeholder)."""
    if val is None:
        return False
    if isinstance(val, str):
        return bool(val.strip()) and not val.startswith("Unnamed: ")
    return True


# =============================================================================
# Row/column trimming
# =============================================================================


def _find_last_data_column(rows: list[tuple]) -> int:
    """Find the 1-based index of the last column with data."""
    if not rows:
        return 0

    max_col = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if _is_cell_non_empty(row[i]):
                max_col = max(max_col, i + 1)
                break
    return max_col


def _find_last_data_row(rows: list[tuple]) -> int:
    """Find the 1-based index of the last row with data."""
    if not rows:
        return 0

    for i in range(len(rows) - 1, -1, -1):
        if any(_is_cell_non_empty(val) for val in rows[i]):
            return i + 1
    return 0


def _is_table_name_row(row: list[Any]) -> bool:
    """Check if row has exactly one meaningful cell (table name pattern)."""
    non_empty = sum(1 for val in row if _is_meaningful_value(val))
    return non_empty == 1 and len(row) > 1


# =============================================================================
# Sheet data extraction
# =============================================================================


def _read_sheet_data(ws: Worksheet) -> tuple[list[dict[str, Any]], list[list[Any]]]:
    """Read sheet data and return (records, all_rows) with trimmed empty cells."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Trim trailing empty rows and columns
    last_row = _find_last_data_row(rows)
    rows = rows[:last_row]
    if not rows:
        return [], []

    last_col = _find_last_data_column(rows)
    rows = [row[:last_col] for row in rows]

    # Generate headers from first row
    headers = [
        (
            f"Unnamed: {i}"
            if val is None or (isinstance(val, str) and not val.strip())
            else str(val)
        )
        for i, val in enumerate(rows[0])
    ]

    # Convert remaining rows to records format
    records: list[dict[str, Any]] = []
    all_rows: list[list[Any]] = [headers]

    for row in rows[1:]:
        record = {
            headers[i]: _get_cell_value(val)
            for i, val in enumerate(row)
            if i < len(headers)
        }
        records.append(record)
        all_rows.append([_get_cell_value(val) for val in row])

    return records, all_rows


def _format_sheet_as_text(all_rows: list[list[Any]]) -> str:
    """Format sheet data as an aligned text table."""
    if not all_rows:
        return ""

    num_cols = max(len(row) for row in all_rows)
    col_widths = [0] * num_cols

    # Format all values and calculate column widths in one pass
    formatted_rows: list[list[str]] = []
    for row in all_rows:
        formatted_row = [
            _format_value_for_display(row[i] if i < len(row) else None)
            for i in range(num_cols)
        ]
        for i, val in enumerate(formatted_row):
            if len(val) > col_widths[i]:
                col_widths[i] = len(val)
        formatted_rows.append(formatted_row)

    return "\n".join(
        " ".join(val.rjust(col_widths[i]) for i, val in enumerate(row))
        for row in formatted_rows
    )


# =============================================================================
# Metadata extraction
# =============================================================================


def _read_metadata(file_like: io.BytesIO) -> XlsxMetadata:
    """Extract document metadata from XLSX core properties."""
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    props = wb.properties

    metadata = XlsxMetadata(
        title=props.title or "",
        description=props.description or "",
        creator=props.creator or "",
        last_modified_by=props.lastModifiedBy or "",
        created=(
            props.created.isoformat()
            if isinstance(props.created, datetime.datetime)
            else ""
        ),
        modified=(
            props.modified.isoformat()
            if isinstance(props.modified, datetime.datetime)
            else ""
        ),
        keywords=props.keywords or "",
        language=props.language or "",
        revision=props.revision,
    )
    wb.close()
    return metadata


# =============================================================================
# Image extraction
# =============================================================================


def _get_content_type(filename: str) -> str:
    """Determine MIME content type from image filename extension."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return _CONTENT_TYPE_MAP.get(ext, "image/unknown")


def _get_image_pixel_dimensions(image_data: bytes) -> tuple[int | None, int | None]:
    """Extract pixel dimensions from common raster formats."""
    if not image_data:
        return None, None

    # PNG
    if image_data.startswith(b"\x89PNG\r\n\x1a\n") and len(image_data) >= 24:
        w = int.from_bytes(image_data[16:20], "big")
        h = int.from_bytes(image_data[20:24], "big")
        return (w or None, h or None)

    # GIF
    if image_data[:6] in (b"GIF87a", b"GIF89a") and len(image_data) >= 10:
        w = int.from_bytes(image_data[6:8], "little")
        h = int.from_bytes(image_data[8:10], "little")
        return (w or None, h or None)

    # BMP
    if image_data[:2] == b"BM" and len(image_data) >= 26:
        w = int.from_bytes(image_data[18:22], "little", signed=True)
        h = int.from_bytes(image_data[22:26], "little", signed=True)
        return (abs(w) or None, abs(h) or None)

    # JPEG
    if image_data.startswith(b"\xff\xd8"):
        i = 2
        size = len(image_data)
        while i + 4 <= size:
            if image_data[i] != 0xFF:
                i += 1
                continue
            marker = image_data[i + 1]
            if marker in (0xD9, 0xDA):
                break
            length = int.from_bytes(image_data[i + 2 : i + 4], "big")
            if length < 2:
                break
            if marker in _JPEG_SOF_MARKERS and i + 2 + length <= size:
                h = int.from_bytes(image_data[i + 5 : i + 7], "big")
                w = int.from_bytes(image_data[i + 7 : i + 9], "big")
                return (w or None, h or None)
            i += 2 + length

    return None, None


def _resolve_drawing_path(target: str) -> str:
    """Normalize drawing relationship targets to ZIP paths."""
    if target.startswith("/"):
        return target[1:]
    if target.startswith(".."):
        return "xl/" + target[3:]
    return "xl/worksheets/" + target


def _resolve_image_path(target: str) -> str:
    """Normalize image relationship targets to ZIP paths."""
    if target.startswith("/"):
        return target[1:]
    return "xl/media/" + target.rsplit("/", 1)[-1]


def _extract_images_from_zip(
    file_like: io.BytesIO, sheet_names: list[str]
) -> dict[int, list[XlsxImage]]:
    """Extract all images from XLSX by parsing the ZIP archive directly."""
    images_by_sheet: dict[int, list[XlsxImage]] = {}
    image_counter = 0

    ctx = OOXMLZipContext(file_like)
    try:
        namelist = ctx.namelist

        # Build mapping of sheet index to drawing file
        sheet_to_drawing: dict[int, str] = {}
        for sheet_idx in range(len(sheet_names)):
            rels_path = f"xl/worksheets/_rels/sheet{sheet_idx + 1}.xml.rels"
            if rels_path not in namelist:
                continue

            rels_root = ctx.read_xml_root(rels_path)
            for rel in parse_relationships(rels_root):
                if "drawing" in rel["type"]:
                    sheet_to_drawing[sheet_idx] = _resolve_drawing_path(rel["target"])
                    break

        # Process each drawing file
        for sheet_idx, drawing_path in sheet_to_drawing.items():
            if drawing_path not in namelist:
                continue

            # Parse drawing relationships to get image file paths
            drawing_rels_path = drawing_path.replace(
                "drawings/", "drawings/_rels/"
            ).replace(".xml", ".xml.rels")

            rid_to_image: dict[str, str] = {}
            if drawing_rels_path in namelist:
                for rel in parse_relationships(ctx.read_xml_root(drawing_rels_path)):
                    if "image" in rel["type"]:
                        rid_to_image[rel["id"]] = _resolve_image_path(rel["target"])

            # Parse the drawing XML to get image metadata
            drawing_root = ctx.read_xml_root(drawing_path)
            sheet_images: list[XlsxImage] = []

            for anchor_type in ANCHOR_TYPES:
                for anchor in drawing_root.iter(anchor_type):
                    pic = anchor.find(XDR_PIC)
                    if pic is None:
                        continue

                    try:
                        # Get dimensions from ext element
                        width, height = 0, 0
                        ext = anchor.find(XDR_EXT)
                        if ext is not None:
                            try:
                                width = int(ext.get("cx", "0")) // EMU_PER_PIXEL
                                height = int(ext.get("cy", "0")) // EMU_PER_PIXEL
                            except ValueError:
                                pass

                        # Get caption and description
                        caption, description = "", ""
                        nvPicPr = pic.find(XDR_NVPICPR)
                        if nvPicPr is not None:
                            cNvPr = nvPicPr.find(XDR_CNVPR)
                            if cNvPr is not None:
                                caption = cNvPr.get("name", "")
                                description = cNvPr.get("descr", "")

                        # Get the blip reference
                        blipFill = pic.find(XDR_BLIPFILL)
                        if blipFill is None:
                            continue

                        blip = blipFill.find(A_BLIP)
                        if blip is None:
                            continue

                        embed_rid = blip.get(R_EMBED, "")
                        if not embed_rid or embed_rid not in rid_to_image:
                            continue

                        image_path = rid_to_image[embed_rid]
                        if image_path not in namelist:
                            continue

                        # Read image data
                        image_bytes = ctx.read_bytes(image_path)
                        filename = image_path.rsplit("/", 1)[-1]

                        if width <= 0 or height <= 0:
                            width, height = _get_image_pixel_dimensions(image_bytes)

                        image_counter += 1
                        sheet_images.append(
                            XlsxImage(
                                image_index=image_counter,
                                sheet_index=sheet_idx,
                                filename=filename,
                                content_type=_get_content_type(filename),
                                data=io.BytesIO(image_bytes),
                                size_bytes=len(image_bytes),
                                width=width,
                                height=height,
                                caption=caption,
                                description=description,
                            )
                        )

                    except Exception as e:
                        logger.debug(f"Failed to extract image from drawing: {e}")

            if sheet_images:
                images_by_sheet[sheet_idx] = sheet_images

    except Exception as e:
        logger.debug(f"Failed to extract images from XLSX: {e}")
    finally:
        ctx.close()

    return images_by_sheet


# =============================================================================
# Content extraction
# =============================================================================


def _read_content(file_like: io.BytesIO) -> list[XlsxSheet]:
    """Read all sheets from XLSX file and extract content."""
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)

    sheet_names = list(wb.sheetnames)
    sheets: list[XlsxSheet] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        records, all_rows = _read_sheet_data(ws)
        text = _format_sheet_as_text(all_rows)

        # Skip first row if it's just a table name
        data_rows = (
            all_rows[1:] if all_rows and _is_table_name_row(all_rows[0]) else all_rows
        )

        sheets.append(
            XlsxSheet(
                name=str(sheet_name),
                data=data_rows,
                text=text,
                images=[],
            )
        )

    wb.close()

    # Extract images by parsing the ZIP archive directly
    images_by_sheet = _extract_images_from_zip(file_like, sheet_names)
    for sheet_idx, sheet_images in images_by_sheet.items():
        if sheet_idx < len(sheets):
            sheets[sheet_idx].images = sheet_images

    return sheets


# =============================================================================
# Main entry point
# =============================================================================


def read_xlsx(
    file_like: io.BytesIO, path: str | None = None
) -> Generator[XlsxContent, Any, None]:
    """
    Extract all relevant content from an Excel .xlsx file.

    Uses a generator pattern for API consistency. XLSX files yield exactly one
    XlsxContent object containing sheets, metadata, and images.
    """
    try:
        file_like.seek(0)
        if is_ooxml_encrypted(file_like):
            raise ExtractionFileEncryptedError(
                "XLSX is encrypted or password-protected"
            )

        validate_zip_bytesio(file_like, source="read_xlsx")

        sheets = _read_content(file_like)
        metadata = _read_metadata(file_like)
        metadata.populate_from_path(path)

        total_rows = sum(len(sheet.data) for sheet in sheets)
        total_images = sum(len(sheet.images) for sheet in sheets)
        logger.info(
            "Extracted XLSX: %d sheets, %d total rows, %d images",
            len(sheets),
            total_rows,
            total_images,
        )

        yield XlsxContent(metadata=metadata, sheets=sheets)
    except ExtractionError:
        raise
    except Exception as exc:
        raise ExtractionFailedError("Failed to extract XLSX file", cause=exc) from exc
